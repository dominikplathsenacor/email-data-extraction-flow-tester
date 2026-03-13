"""REST roundtrip integration tests."""

from __future__ import annotations

import json
import threading
from contextlib import contextmanager
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import ClassVar

from openpyxl import load_workbook
from simple_e2e_tester.configuration.loader import load_configuration
from simple_e2e_tester.run_execution.run_contracts import RunRequest
from simple_e2e_tester.run_execution.validation_run_use_case import (
    execute_email_kafka_validation_run,
)
from simple_e2e_tester.schema_management import flatten_schema, load_schema_document
from simple_e2e_tester.template_generation import (
    TEMPLATE_SHEET_NAME,
    generate_template_workbook,
)


class _RestHandler(BaseHTTPRequestHandler):
    """In-test REST handler that captures request payloads."""

    response_payload: ClassVar[dict[str, object]] = {}
    captured_requests: ClassVar[list[dict[str, object]]] = []

    def do_POST(self) -> None:  # pylint: disable=invalid-name
        content_length = int(self.headers.get("Content-Length", "0"))
        raw_body = self.rfile.read(content_length)
        payload = json.loads(raw_body.decode("utf-8"))
        _RestHandler.captured_requests.append(payload)
        encoded = json.dumps(_RestHandler.response_payload).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)

    def log_message(
        self, format: str, *args
    ) -> None:  # pylint: disable=redefined-builtin
        del format, args


@contextmanager
def _run_fake_rest_server(response_payload: dict[str, object]):
    """Run a local HTTP server for the test lifetime."""
    _RestHandler.response_payload = response_payload
    _RestHandler.captured_requests = []
    server = ThreadingHTTPServer(("127.0.0.1", 0), _RestHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        yield f"http://127.0.0.1:{server.server_port}", _RestHandler.captured_requests
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=2.0)


def test_given_rest_mode_when_run_executes_then_real_http_roundtrip_is_validated(
    tmp_path: Path,
) -> None:
    response_payload = {
        "sender": "sender@example.com",
        "subject": "Subject-1",
        "score": 1.58,
    }
    with _run_fake_rest_server(response_payload) as (base_url, captured_requests):
        config_path = _write_rest_config(tmp_path, base_url)
        template_path = _write_template(tmp_path, config_path)
        output_dir = tmp_path / "results"

        outcome = execute_email_kafka_validation_run(
            RunRequest(
                config_path=str(config_path),
                input_path=str(template_path),
                output_dir=str(output_dir),
                dry_run=False,
            )
        )

    assert outcome.sent_ok == 1
    assert len(captured_requests) == 1
    assert captured_requests[0]["dok_text"] == "Body value from testcase"
    assert captured_requests[0]["emailbetreff"] == "Subject-1"
    assert captured_requests[0]["emailabsender"] == "sender@example.com"

    workbook = load_workbook(outcome.output_path)
    testcase_sheet = workbook[TEMPLATE_SHEET_NAME]
    assert testcase_sheet.cell(row=3, column=testcase_sheet.max_column).value == "OK"

    run_info_sheet = workbook["RunInfo"]
    run_info = {
        row[0].value: row[1].value for row in run_info_sheet.iter_rows(min_row=2)
    }
    assert run_info["kafka_topic"] == "REST_DIRECT"


def _write_rest_config(tmp_path: Path, base_url: str) -> Path:
    schema = {
        "type": "object",
        "properties": {
            "sender": {"type": "string"},
            "subject": {"type": "string"},
            "score": {"type": "number"},
        },
    }
    config = {
        "transport": {"mode": "rest"},
        "schema": {"rest_response": {"json_schema": {"inline": json.dumps(schema)}}},
        "matching": {"from_field": "sender", "subject_field": "subject"},
        "smtp": {"host": "smtp.example.com", "port": 25},
        "mail": {"to_address": "qa@example.com"},
        "rest": {
            "base_url": base_url,
            "path": "/extract",
            "method": "POST",
            "defaults": {
                "ag": "AG-1",
                "dokart": "DOKART-1",
                "dokrefuid": "DOKREF-1",
                "eingangsdatum": "2026-01-01-00.00.00.000000",
                "flowid": "FLOW-1",
                "ordnungsbegriff": "ORD-1",
                "referenztyp": "EM",
            },
        },
    }
    config_path = tmp_path / "config.json"
    config_path.write_text(json.dumps(config), encoding="utf-8")
    return config_path


def _write_template(tmp_path: Path, config_path: Path) -> Path:
    configuration = load_configuration(config_path)
    fields = flatten_schema(load_schema_document(configuration.schema))
    template_path = tmp_path / "template.xlsx"
    generate_template_workbook(configuration.schema, fields, template_path)

    workbook = load_workbook(template_path)
    sheet = workbook[TEMPLATE_SHEET_NAME]
    headers = {
        sheet.cell(row=2, column=col).value: col
        for col in range(1, sheet.max_column + 1)
    }
    sheet.cell(row=3, column=headers["ID"]).value = "TC-REST-1"
    sheet.cell(row=3, column=headers["FROM"]).value = "sender@example.com"
    sheet.cell(row=3, column=headers["SUBJECT"]).value = "Subject-1"
    sheet.cell(row=3, column=headers["BODY"]).value = "Body value from testcase"
    sheet.cell(row=3, column=headers["score"]).value = 1.58
    workbook.save(template_path)
    return template_path
